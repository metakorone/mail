import tkinter as tk
from tkinter import ttk
from tkinter.constants import SE
from tkinter.scrolledtext import ScrolledText
from tkinter import messagebox
from tkinter import filedialog
import openpyxl
import win32com.client
import os
import datetime
import jpbizday
import shutil
import json
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.mime.text import MIMEText
from email.utils import formatdate
import subprocess
from time import sleep
import numpy as np

class MainWindow(tk.Frame):
    def __init__(self, master):
        super().__init__(master)
        #中身に応じて自動調節してくれた！！
        self.pack(fill="both", expand=True)
        #ウィンドウの詳細設定
        self.master.title("出校記録提出")
        self.main_window()
        if not os.path.isdir("setting"):
            os.mkdir("setting")
    #メインウィンドウのウィジェット配置
    def main_window(self):
        #いつの週の出向記録を作るかのボタン
        def week_button(text0, row0, column0, func):
            self.button_week = ttk.Button(self)
            self.button_week.configure(text=text0)
            self.button_week.bind("<Button-1>", func) #do not forget to add self!
            self.button_week.grid(row=row0, column=column0, padx=5, pady=20)

        #コールバック用の関数
        def callback(event):
            #押されたボタンによってどのウィンドウを開くか振り分け．switchほしい
            #また，初回起動時setting.jsonが存在しないのでどのボタンを押してもsettingwindowを開くようにしている．
            #aは設定ファイルが全部存在しているか確認するための変数
            a=0
            if not os.path.exists('setting/template.json'):
                self.creating_window("template")
                a=1
            if not os.path.exists('setting/send_setting.json'):
                self.creating_window("send")
                a=1
            if not os.path.exists('setting/pdf.json'):
                self.creating_window("pdf")
                a=1
            if not os.path.exists('setting/setting.json'):
                self.creating_window("setting")
                a=1
            if a == 0:
                if event.widget["text"] == "今週":
                    self.creating_window("record1")
                elif event.widget["text"] == "先週":
                    self.creating_window("record3")
                elif event.widget["text"] == "来週":
                    self.creating_window("record2")

        #メインウィンドウのウィジェット配置
        self.label = ttk.Label(self)
        self.label.configure(text="いつの週の出校記録を作りますか？")
        self.label.grid(row=0, columnspan=3, padx=5, pady=20)
        week_button("先週", 1 ,0, callback)
        week_button("今週", 1, 1, callback)
        week_button("来週", 1, 2, callback)

        #メニューバー
        #個人設定を押したとき用
        def setting():
            self.creating_window("setting")

        #テンプレートを編集するを押したとき用
        def setting_temp():
            self.creating_window("template")

        #送信設定を押したとき用
        def setting_send():
            self.creating_window("send")

        #pdfを押したとき用
        def setting_pdf():
            self.creating_window("pdf")

        menubar = tk.Menu(self)
        self.master.configure(menu=menubar)
        #メニューバーに設定を作成
        menu_setting = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="設定", menu=menu_setting)
        #設定に個人設定，送信設定を追加
        menu_setting.add_command(label='個人設定', command=setting)
        menu_setting.add_command(label='送信設定', command=setting_send)
        menu_setting.add_command(label='テンプレートを編集する', command=setting_temp)
        menu_setting.add_command(label='pdfリーダーの設定', command=setting_pdf)


    #settingwindowやtemplatewindowを開く処理を一つの関数として纏めておくことでこれから昨日を追加したときにやりやすいようにしておく．
    def creating_window(self, what):
        if what == "setting":
            sub = tk.Toplevel(self.master)
            app2 = SettingWindow(master = sub)
            app2.wait_window()
        elif what == "template":
            sub2 = tk.Toplevel(self.master)
            app3 = TemplateWindow(master = sub2)
            app3.wait_window()
        elif what == "send":
            sub3 = tk.Toplevel(self.master)
            app4 = SendSettingWindow(master = sub3)
            app4.wait_window()
        elif what == "pdf":
            sub5 = tk.Toplevel(self.master)
            app6 = PdfSetting(master = sub5)
            app6.wait_window()
        elif what == "record1":
            sub4 = tk.Toplevel(self.master)
            app5 = SendRecord1(master = sub4)
            app5.wait_window()
        elif what == "record2":
            sub4 = tk.Toplevel(self.master)
            app5 = SendRecord3(master = sub4)
            app5.wait_window()
        elif what == "record3":
            sub4 = tk.Toplevel(self.master)
            app5 = SendRecord2(master = sub4)
            app5.wait_window()

class SettingWindow(tk.Frame):
    def __init__(self, master):
        super().__init__(master)
        #中身に応じて自動調節してくれた！！
        self.pack(fill="both", expand=True)
        #ウィンドウの詳細設定
        self.master.title("設定")
        self.setting_window()
    #settingwindowのウィジェット配置，入力したデータを設定ファイルに出力する関数（本来はこれを外に出したかったがうまく動かなくなってしまったので内包）
    def setting_window(self):

        #コールバック用の関数
        def callback(event):
            save_json()
            self.master.destroy()

        #設定ファイルに出力する関数(destroyをこの関数の末尾につけるもうまくいかず)
        def save_json():
            value1 = editBoxFName.get()
            value2 = editBoxName.get()
            value3 = editBoxs_id.get()
            with open('setting/setting.json', 'w') as f:
                json.dump({"name": [value1,value2], "s_id": f"{value3}"}, f, indent=4)

        #既に設定ファイルが存在したときに，入力ボックスに設定内容を表示するために設定ファイルから情報を読み取り変数に格納
        if os.path.exists('setting/setting.json') == True:
            with open('setting/setting.json', 'r') as f:
                json_dic = json.load(f)
        else:
            json_dic = {"name": ["",""], "s_id": ""}

        #複数のラベルを用いるのでモジュール化
        def labels(text0, row0, column0, columnspan0):
            self.label = ttk.Label(self)
            self.label.configure(text=text0)
            self.label.grid(row=row0, column=column0, columnspan=columnspan0, padx=5, pady=10)

        labels("設定ファイルを作成します\n下記の必要事項を入力後，\n保存ボタンを押しウィンドウを閉じてください\nその後もう一度作りたい週のボタンを押して下さい．", 0, 0, 2)
        labels("氏名", 1, 0, 2)
        labels("姓", 2, 0, 1)
        labels("名", 2, 1, 1)
        labels("学籍番号", 4, 0, 2)
        # エントリーボックスをモジュールかしようと思ったが名前をつけることができなかったため断念
        editBoxFName = ttk.Entry(self)
        editBoxFName.insert(tk.END, json_dic.get("name")[0])
        editBoxFName.grid(row=3, column=0, columnspan=1, padx=5, pady=10)
        editBoxName = ttk.Entry(self)
        editBoxName.insert(tk.END, json_dic.get("name")[1])
        editBoxName.grid(row=3, column=1, columnspan=1, padx=5)
        editBoxs_id = ttk.Entry(self, width=30)
        editBoxs_id.insert(tk.END, json_dic.get("s_id"))
        editBoxs_id.grid(row=5,columnspan=2)
        #保存ボタン
        button_sub = ttk.Button(self, text="保存")
        button_sub.bind("<Button-1>", callback)
        button_sub.grid(row=6, column=1, padx=5, pady=20,sticky=tk.SE)

class TemplateWindow(tk.Frame):
    def __init__(self, master):
        super().__init__(master)
        #中身に応じて自動調節してくれた！！
        self.pack(fill="both", expand=True)
        #ウィンドウの詳細設定
        self.master.title("テンプレート")
        self.template_window()

    def template_window(self):
        #コールバック用の関数
        def callback(event):
            save_json()
            self.master.destroy()
        #テンプレートを保存する関数
        def save_json():
            value1 = textbox.get("1.0", "end")
            with open('setting/template.json', 'w') as f:
                json.dump({"text": f"{value1}"}, f, indent=4)

        #既に設定ファイルが存在したときに，入力ボックスに設定内容を表示するために設定ファイルから情報を読み取り変数に格納
        if os.path.exists('setting/template.json') == True:
            with open('setting/template.json', 'r') as f:
                json_dic = json.load(f)
        else:
            json_dic = {"text": "宮本先生，後藤先生，松村様\n\n{now_month}/{now_day}週の出校記録です．よろしくお願いします\n\n{name[0]}\n\n"}

        #説明文
        self.label = ttk.Label(self)
        self.label.configure(text="メール本文のテンプレートを作成します\nこだわりなければ何も変えずに保存で大丈夫です\n特別な変数\n{now_month}:選択した週の月曜日の月\n{now_day}:選択した週の月曜日の日\n{name[0]}:設定した姓")
        self.label.pack()
        #文章ボックスの作成
        textbox = ScrolledText(self, font=("", 13), height=10, width=40)
        textbox.pack()
        textbox.insert(tk.END, json_dic.get("text"))
        #保存ボタン
        button_sub = ttk.Button(self, text="保存")
        button_sub.bind("<Button-1>", callback)
        button_sub.pack()

class SendSettingWindow(tk.Frame):
    def __init__(self, master):
        super().__init__(master)
        #中身に応じて自動調節してくれた！！
        self.pack(fill="both", expand=True)
        #ウィンドウの詳細設定
        self.master.title("送信設定")
        self.send_setting_window()

    #設定があっているかテストするためにメール送る用（自分にメールを送る）
    def testmail(self):
        #設定ファイル読み込み
        with open('setting/send_setting.json', 'r') as f:
            json_dic = json.load(f)
        # 送受信先
        to_addr = json_dic.get("mail")
        from_addr = json_dic.get("mail")

        msg = MIMEText('テスト', "plain", 'utf-8')
        msg['Subject'] = "【出校記録自動提出】テスト"
        msg['From'] = from_addr
        msg['To'] = to_addr
        msg['Date'] = formatdate()

        with smtplib.SMTP_SSL(host=json_dic.get("server"), port=json_dic.get("port")) as smtp:
            smtp.login(json_dic.get("id"), json_dic.get("pass"))
            smtp.send_message(msg)
            smtp.quit()

    #テストメールを送信するかどうか確認のダイアログ
    def check(self):
        Messagebox = tk.messagebox.askquestion('テスト','テストメールを送信しますか？')
        if Messagebox == 'yes': #If関数
            self.testmail()

    def send_setting_window(self):
        #送信設定のラベル用モジュール
        def labels(text0):
            self.label = ttk.Label(self)
            self.label.configure(text = text0)
            self.label.pack()

        #コールバック用の関数
        def callback(event):
            save_json()
            self.check()
            self.master.destroy()

        #設定ファイルに出力する関数(destroyをこの関数の末尾につけるもうまくいかず)
        def save_json():
            value1 = editBoxServer.get()
            value2 = editBoxPort.get()
            value3 = editBoxID.get()
            value4 = editBoxPass.get()
            value5 = editBoxMail.get()
            with open('setting/send_setting.json', 'w') as f:
                json.dump({"server": value1, "port": value2, "id": value3, "pass": value4, "mail": value5}, f, indent=4)

        #既に設定ファイルが存在したときに，入力ボックスに設定内容を表示するために設定ファイルから情報を読み取り変数に格納
        if os.path.exists('setting/send_setting.json') == True:
            with open('setting/send_setting.json', 'r') as f:
                json_dic = json.load(f)
        else:
            json_dic = {"server": "", "port": "", "id": "", "pass": "","mail": ""}

        #ウィジェットの配置
        labels("送信サーバー")
        editBoxServer = ttk.Entry(self)
        editBoxServer.insert(tk.END, json_dic.get("server"))
        editBoxServer.pack()
        labels("ポート")
        editBoxPort = ttk.Entry(self)
        editBoxPort.insert(tk.END, json_dic.get("port"))
        editBoxPort.pack()
        labels("ID")
        editBoxID = ttk.Entry(self)
        editBoxID.insert(tk.END, json_dic.get("id"))
        editBoxID.pack()
        labels("Password")
        editBoxPass = ttk.Entry(self)
        editBoxPass.insert(tk.END, json_dic.get("pass"))
        editBoxPass.pack()
        labels("メールアドレス")
        editBoxMail = ttk.Entry(self)
        editBoxMail.insert(tk.END, json_dic.get("mail"))
        editBoxMail.pack()
        #保存ボタン
        button_sub = ttk.Button(self, text="保存")
        button_sub.bind("<Button-1>", callback)
        button_sub.pack(pady=10)

class PdfSetting(tk.Frame):
    def __init__(self, master):
        super().__init__(master)
        #中身に応じて自動調節してくれた！！
        self.pack(fill="both", expand=True)
        #ウィンドウの詳細設定
        self.master.title("pdfリーダー")
        self.pdfsetting()

    def pdfsetting(self):
        #コールバック用の関数
        def callback(event):
            save_json()
            self.master.destroy()

        #コールバック用の関数
        def callback2(event):
            file_browser()

        #テンプレートを保存する関数
        def save_json():
            value1 = editBoxPath.get()
            with open('setting/pdf.json', 'w') as f:
                json.dump({"path": f"{value1}"}, f, indent=4)

        def file_browser():
            typ = [('実行ファイル','*.exe')]
            dir = 'C:\Program Files'
            fle = filedialog.askopenfilename(filetypes = typ, initialdir = dir)
            editBoxPath.insert(tk.END, fle)

        #既に設定ファイルが存在したときに，入力ボックスに設定内容を表示するために設定ファイルから情報を読み取り変数に格納
        if os.path.exists('setting/pdf.json') == True:
            with open('setting/pdf.json', 'r') as f:
                json_dic = json.load(f)
        else:
            json_dic = {"path": ""}

        #説明文
        self.label = ttk.Label(self)
        self.label.configure(text="pdfを読み取れるアプリの実行ファイルのパスを指定してください")
        self.label.pack()
        #文章ボックスの作成
        editBoxPath = ttk.Entry(self)
        editBoxPath.insert(tk.END, json_dic.get("path"))
        editBoxPath.pack()
        #開くボタン
        button_sub = ttk.Button(self, text="開く...")
        button_sub.bind("<Button-1>", callback2)
        button_sub.pack()
        #保存ボタン
        button_sub = ttk.Button(self, text="保存")
        button_sub.bind("<Button-1>", callback)
        button_sub.pack()

class SendRecord(tk.Frame):
    #選んだ週の月曜日の日付を取得
    def getAnyday(self, day):
            m_then = str(datetime.datetime.now() - datetime.timedelta(days = datetime.date.today().weekday() + day))
            now_year = m_then[0:4]
            now_month = m_then[5:7]
            now_day = m_then[8:10]
            return now_year, now_month, now_day

    def check(self, msg, json_dic):
        result = tk.messagebox.askquestion('出校記録送信',f'出校記録を送信しますか？\n{msg}')
        if result == 'yes': #If関数
            self.sendmail(msg, json_dic)

    def sendmail(self, msg, json_dic):
        with smtplib.SMTP_SSL(host=json_dic.get("server"), port=json_dic.get("port")) as smtp:
            smtp.login(json_dic.get("id"), json_dic.get("pass"))
            smtp.send_message(msg)
            smtp.quit()

    def record(self, var):
        with open('setting/setting.json', 'r') as f:
            json_dic = json.load(f)
        name = ''.join(json_dic.get("name"))
        s_id = json_dic.get("s_id")

        #path(出校記録のアーカイブを保存したいディレクトリを絶対パスで指定してください)
        curPath = os.getcwd()

        #今週の月曜日からday日後の年月日を取得
        def getanyday(day):
            m_then = str(datetime .datetime.now() - datetime.timedelta(days = datetime.date.today().weekday() + day))
            global m_then_year, m_then_month, m_then_day
            m_then_year = m_then[0:4]
            m_then_month = m_then[5:7]
            m_then_day = m_then[8:10]

        #拾い物
        def xl2pdf(excel_path, pdf_path):
            excel = win32com.client.Dispatch("Excel.Application")    # Excelの起動
            file = excel.Workbooks.Open(excel_path)    # Excelファイルを開く
            file.WorkSheets("入力シート").Activate()    # Sheetをシート名でアクティベイト
            file.ActiveSheet.ExportAsFixedFormat(0, pdf_path)    # PDFに変換
            file.Close()    # 開いたエクセルを閉じる
            excel.Quit()    # Excelを終了

        i_week = -(self.week * 7)
        getanyday(i_week)
        #日付で分類したフォルダの作成&エクセルファイルのコピーとリネーム
        c_path = curPath + f'\\{m_then_year}年度\\{m_then_month}月\\{m_then_month}月{m_then_day}日週分'
        c_xl_path = c_path + f'\\{name}_{s_id}_{m_then_month}{m_then_day}.xlsx'
        c_pdf_path = c_path + f'\\{name}_{s_id}_{m_then_month}{m_then_day}.pdf'
        os.makedirs(c_path, exist_ok=True)
        shutil.copyfile(curPath + '\\origin.xlsx', c_xl_path)
        #エクセルのファイルごにょごにょ
        wb = openpyxl.load_workbook(c_xl_path)
        ws = wb.active
        #こういう変数の使い方をしたくない
        count = 0
        #日付入力
        for i in range(0,-7,-1):
            getanyday((i+ i_week))
            #チェックボックスにチェックを入れた曜日のみ出力
            bools = tk.BooleanVar()
            bools.set(True)
            #営業日判定、土日ではない、祝日ではない、1\\1から1\\3ではないとき       を入校日に入れてるらしい
            if var[-i].get():
                ws.cell(row=count+6, column=2, value=f'{int(m_then_month)}月{int(m_then_day)}日')
                count =  count + 1
        #祝日は入力させないので余ってしまったセルをワイプしてる
        for i in range(5):
            if (count +i) < 6:
                for j in range(22):
                    ws.cell(row=count+i+6, column=j+1, value='')
        wb.save(c_xl_path)
        xl2pdf(c_xl_path, c_pdf_path)
        with open('setting/pdf.json', 'r') as f:
            json_pdf = json.load(f)
        pdf = subprocess.Popen([json_pdf.get("path"), c_pdf_path], shell=False)

    #メールを送信する
    def mail(self, m_then_year, m_then_month, m_then_day):
        #設定ファイル読み込み
        with open('setting/send_setting.json', 'r') as f:
            json_dic = json.load(f)
        with open('setting/template.json', 'r') as f:
            json_tem = json.load(f)
        with open('setting/setting.json', 'r') as f:
            json_set = json.load(f)
        # 送受信先
        from_addr = json_dic.get("mail")

        now_month = m_then_year
        now_day = m_then_day
        name = json_set.get("name")

        msg = MIMEMultipart()
        msg['Subject'] = f"出校記録_{json_set.get('name')[0]}"
        msg['From'] = from_addr
        msg['To'] = from_addr
        msg['Date'] = formatdate()
        body = MIMEText(f'宮本先生，後藤先生，松村様\n\n{m_then_month}/{m_then_day}週の出校記録です．よろしくお願いします\n\n{name[0]}\n\n', "plain", 'utf-8')
        msg.attach(body)

        # ファイルを添付
        c_path = os.getcwd() + f'\\{m_then_year}年度\\{m_then_month}月\\{m_then_month}月{m_then_day}日週分'
        attach = os.listdir(c_path)
        file = attach[0]
        fname = os.path.basename(file)
        file = os.getcwd() + f'\\{m_then_year}年度\\{m_then_month}月\\{m_then_month}月{m_then_day}日週分\\{fname}'
        with open(file, "rb") as f:
            part = MIMEApplication(
                f.read(),
                Name=fname
            )
        part['Content-Disposition'] = 'attachment; filename="%s"' % fname
        msg.attach(part)

        file = attach[1]
        fname = os.path.basename(file)
        file = os.getcwd() + f'\\{m_then_year}年度\\{m_then_month}月\\{m_then_month}月{m_then_day}日週分\\{fname}'
        with open(file, "rb") as f:
            part1 = MIMEApplication(
                f.read(),
                Name=fname
            )
        part1['Content-Disposition'] = 'attachment; filename="%s"' % fname
        msg.attach(part1)

        self.check(msg, json_dic)

    def send_record(self):
        #コールバック用の関数
        def callback(event):
            self.record(self.var)
            m_then = self.getAnyday(-7*self.week)
            now_year = m_then[0]
            now_month = m_then[1]
            now_day = m_then[2]
            self.mail(now_year, now_month, now_day)
            self.master.destroy()

        #一週間分のチェックボックスを作る
        list_chk = {}
        self.var = {}
        for i in range(7):
            m_then = self.getAnyday(-7*self.week - i)
            now_year = m_then[0]
            now_month = m_then[1]
            now_day = m_then[2]
            if i == 0:
                day_of_week = "月"
            elif i == 1:
                day_of_week = "火"
            elif i == 2:
                day_of_week = "水"
            elif i == 3:
                day_of_week = "木"
            elif i == 4:
                day_of_week = "金"
            elif i == 5:
                day_of_week = "土"
            elif i == 6:
                day_of_week = "日"
            #営業日をチェック済みにしておく
            self.var[i] = tk.BooleanVar()
            self.var[i].set(jpbizday.is_bizday(datetime.date(int(now_year), int(now_month), int(now_day))))
            # チェックボタン作成
            list_chk[i] = ttk.Checkbutton(self, text=f"{day_of_week}曜日({now_month}/{now_day})", variable=self.var[i])
            list_chk[i].pack()

        button = ttk.Button(self, text="送信")
        button.bind("<Button-1>", callback)
        button.pack(pady=10)

#weeksをインスタンス化の時に渡したかったがtkinterの使用を理解できていなかったためクラスを分割した
class SendRecord1(SendRecord):
    def __init__(self, master):
        self.week = 0
        super().__init__(master)
        #中身に応じて自動調節してくれた！！
        self.pack(fill="both", expand=True)
        #ウィンドウの詳細設定
        self.master.title("出校記録送信")
        self.send_record()

class SendRecord2(SendRecord):
    def __init__(self, master):
        self.week = -1
        super().__init__(master)
        #中身に応じて自動調節してくれた！！
        self.pack(fill="both", expand=True)
        #ウィンドウの詳細設定
        self.master.title("出校記録送信")
        self.send_record()

class SendRecord3(SendRecord):
    def __init__(self, master):
        self.week = 1
        super().__init__(master)
        #中身に応じて自動調節してくれた！！
        self.pack(fill="both", expand=True)
        #ウィンドウの詳細設定
        self.master.title("出校記録送信")
        self.send_record()


def main():
    root = tk.Tk()
    app = MainWindow(master = root)
    app.mainloop()

if __name__ == "__main__":
    main()